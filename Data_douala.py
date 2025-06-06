import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Clé API WeatherAPI.com
API_KEY = "de330782dc16418397b172524250605"
CITY = "Douala"
URL = f"http://api.weatherapi.com/v1/current.json?key={API_KEY}&q={CITY}&aqi=no"

# Nom du fichier Excel
fichier_excel = "meteo_douala5.xlsx"

# Initialise le fichier s’il n’existe pas
if not os.path.exists(fichier_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Météo"
    ws.append([
        "Date et Heure", "ID Station", "Latitude", "Longitude",
        "Ville", "Région", "Pays", "Description météo",
        "Température (°C)", "Pression (hPa)", "Humidité (%)",
        "Vitesse du vent (m/s)", "Précipitation (mm)"
    ])
    wb.save(fichier_excel)

# Fonction pour géocoder à partir de coordonnées (idem)
def get_adresse_reverse(lat, lon):
    try:
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
        headers = {'User-Agent': 'MeteoApp'}
        res = requests.get(url, headers=headers)
        if res.status_code == 200:
            data = res.json().get("address", {})
            ville = data.get("city") or data.get("town") or data.get("village") or ""
            region = data.get("state") or ""
            pays = data.get("country") or ""
            return ville, region, pays
    except Exception as e:
        print("🌍 Exception géocodage :", e)
    return "", "", ""

# Fonction d’ajout des données météo
def ajouter_donnees_excel(data):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # WeatherAPI ne fournit pas d'ID station global équivalent OpenWeather
    station_id = data.get('location', {}).get('tz_id', '')
    lat = data['location']['lat']
    lon = data['location']['lon']
    ville, region, pays = get_adresse_reverse(lat, lon)
    description = data['current']['condition']['text']
    temperature = data['current']['temp_c']
    pression = data['current']['pressure_mb']
    humidite = data['current']['humidity']
    vent = data['current']['wind_kph'] / 3.6  # convertir km/h → m/s
    # WeatherAPI renvoie pluie sur 1h et 24h
    precipitation = data['current'].get('precip_mm', 0.0)

    wb = load_workbook(fichier_excel)
    ws = wb.active
    ws.append([
        now, station_id, lat, lon,
        ville, region, pays, description,
        temperature, pression, humidite, vent, precipitation
    ])
    wb.save(fichier_excel)
    print(f"✅ Données ajoutées à {fichier_excel} à {now}")

# Boucle toutes les 10 minutes
if __name__ == "__main__":
    while True:
        try:
            reponse = requests.get(URL, timeout=10)
            if reponse.status_code == 200:
                donnees = reponse.json()
                ajouter_donnees_excel(donnees)
            else:
                print("⚠️ Erreur API météo :", reponse.status_code, reponse.text)
        except Exception as e:
            print("❌ Erreur :", e)
        # 600 secondes = 10 minutes
        time.sleep(600)
