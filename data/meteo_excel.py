import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Clé API météo
API_KEY = "86d9c4e3060dfe17573901cb4526e82a"
CITY = "Douala"
URL = f"https://api.openweathermap.org/data/2.5/weather?q={CITY}&appid={API_KEY}&units=metric"

# Nom du fichier Excel
fichier_excel = "meteo_douala4.xlsx"

# Initialise le fichier s’il n’existe pas
if not os.path.exists(fichier_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Météo"
    ws.append([
        "Date et Heure", "ID Station", "Latitude", "Longitude",
        "Ville", "Région", "Pays", "Description météo",
        "Température (°C)", "Pression (hPa)", "Humidité (%)",
        "Vitesse du vent (m/s)", "Précipitation (mm)"
    ])
    wb.save(fichier_excel)

# Fonction pour géocoder à partir de coordonnées
def get_adresse_reverse(lat, lon):
    try:
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
        headers = {'User-Agent': 'MeteoApp'}
        res = requests.get(url, headers=headers)
        if res.status_code == 200:
            data = res.json()
            address = data.get("address", {})
            ville = address.get("city") or address.get("town") or address.get("village") or ""
            region = address.get("state") or ""
            pays = address.get("country") or ""
            return ville, region, pays
        else:
            print(f"🌍 Erreur Nominatim : {res.status_code}")
            return "", "", ""
    except Exception as e:
        print("🌍 Exception géocodage :", e)
        return "", "", ""

# Fonction d’ajout des données météo
def ajouter_donnees_excel(data):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    station_id = data.get('id', '')
    lat = data['coord']['lat']
    lon = data['coord']['lon']
    ville, region, pays = get_adresse_reverse(lat, lon)
    description = data['weather'][0]['description']
    temperature = data['main']['temp']
    pression = data['main']['pressure']
    humidite = data['main']['humidity']
    vent = data['wind']['speed']
    precipitation = data.get('rain', {}).get('1h', 0.0)

    wb = load_workbook(fichier_excel)
    ws = wb.active
    ws.append([
        now, station_id, lat, lon,
        ville, region, pays, description,
        temperature, pression, humidite, vent, precipitation
    ])
    wb.save(fichier_excel)
    print(f"✅ Données ajoutées à {fichier_excel} à {now}")

# Boucle toutes les 10 minutes
while True:
    try:
        reponse = requests.get(URL)
        if reponse.status_code == 200:
            donnees = reponse.json()
            ajouter_donnees_excel(donnees)
        else:
            print("⚠️ Erreur API météo :", reponse.status_code)
    except Exception as e:
        print("❌ Erreur :", e)

    time.sleep(600)
