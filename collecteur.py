import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Clé API OpenWeatherMap
# en haut de collecte/collecteur.py
API_KEY = "86d9c4e3060dfe17573901cb4526e82a"

# Liste des villes à collecter
VILLES = ["Douala"]
# Fichier Excel de sortie
FICHIER_EXCEL = os.path.join(os.path.dirname(__file__), "..", "data", "meteo_donnees.xlsx")
INTERVALLE = 10 * 60  # 10 minutes

# Initialisation du fichier Excel
if not os.path.exists(FICHIER_EXCEL):
    wb = Workbook()
    for ville in VILLES:
        ws = wb.create_sheet(title=ville)
        ws.append([
            "Date et Heure", "Temperature", "Pression", "Humidite", "Vent"
        ])
    # Supprimer la feuille par défaut si vide
    default = wb['Sheet']
    wb.remove(default)
    os.makedirs(os.path.dirname(FICHIER_EXCEL), exist_ok=True)
    wb.save(FICHIER_EXCEL)


def collecter():
    wb = load_workbook(FICHIER_EXCEL)
    for ville in VILLES:
        url = f"https://api.openweathermap.org/data/2.5/weather?q={ville}&appid={API_KEY}&units=metric"
        resp = requests.get(url)
        if resp.status_code == 200:
            data = resp.json()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            temp = data['main']['temp']
            pres = data['main']['pressure']
            hum = data['main']['humidity']
            vent = data['wind']['speed']
            ws = wb[ville]
            ws.append([now, temp, pres, hum, vent])
            print(f"✅ {ville} : données ajoutées à {now}")
        else:
            print(f"⚠️ Erreur pour {ville} : code {resp.status_code}")
    wb.save(FICHIER_EXCEL)


if __name__ == '__main__':
    while True:
        collecter()
        time.sleep(INTERVALLE)