import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Votre clé WeatherAPI
API_KEY = "de330782dc16418397b172524250605"
COORD   = "4.045,9.707"   # Port de Douala

URL = f"http://api.weatherapi.com/v1/current.json?key={API_KEY}&q={COORD}&aqi=no"
fichier_excel = "meteo_port_douala.xlsx"

# Création du fichier Excel si besoin
if not os.path.exists(fichier_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Météo"
    ws.append([
        "Date et Heure", "ID Station", "Latitude", "Longitude",
        "Ville", "Région", "Pays", "Description météo",
        "Température (°C)", "Pression (hPa)", "Humidité (%)",
        "Vitesse du vent (m/s)", "Précipitation (mm)"
    ])
    wb.save(fichier_excel)

def get_adresse_reverse(lat, lon):
    try:
        res = requests.get(
            f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json",
            headers={'User-Agent': 'MeteoApp'}
        )
        addr = res.json().get("address", {})
        return (
            addr.get("city") or addr.get("town") or addr.get("village") or "",
            addr.get("state") or "",
            addr.get("country") or ""
        )
    except:
        return "", "", ""

def ajouter_donnees_excel(data):
    now         = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    station_id  = data['location'].get('tz_id', '')
    lat, lon    = data['location']['lat'], data['location']['lon']
    ville, region, pays = get_adresse_reverse(lat, lon)
    cur         = data['current']
    desc        = cur['condition']['text']
    temp        = cur['temp_c']
    pres        = cur['pressure_mb']
    hum         = cur['humidity']
    wind        = cur['wind_kph'] / 3.6
    precip      = cur.get('precip_mm', 0.0)

    wb = load_workbook(fichier_excel)
    ws = wb.active
    ws.append([
        now, station_id, lat, lon,
        ville, region, pays, desc,
        temp, pres, hum, wind, precip
    ])
    wb.save(fichier_excel)
    print(f"✅ Données ajoutées à {fichier_excel} à {now}")

if __name__ == "__main__":
    while True:
        try:
            r = requests.get(URL, timeout=10)
            if r.ok:
                ajouter_donnees_excel(r.json())
            else:
                print("⚠️ Erreur API :", r.status_code)
        except Exception as e:
            print("❌ Erreur :", e)
        time.sleep(60)  # 10 minutes
