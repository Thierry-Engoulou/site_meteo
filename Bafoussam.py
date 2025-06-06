import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Votre clé WeatherAPI
API_KEY = "de330782dc16418397b172524250605"

# Ville ciblée
CITY = "Bafoussam"

# URL WeatherAPI pour les données actuelles
URL = f"http://api.weatherapi.com/v1/current.json?key={API_KEY}&q={CITY}&aqi=no"

# Nom du fichier Excel de sortie
fichier_excel = "meteo_bafoussam.xlsx"

# Initialise le fichier s’il n’existe pas
if not os.path.exists(fichier_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Météo"
    ws.append([
        "Date et Heure", "ID Station", "Latitude", "Longitude",
        "Ville", "Région", "Pays", "Description météo",
        "Température (°C)", "Pression (hPa)", "Humidité (%)",
        "Vitesse du vent (m/s)", "Précipitation (mm)"
    ])
    wb.save(fichier_excel)

def get_adresse_reverse(lat, lon):
    """
    Retourne ville, région, pays via OpenStreetMap/Nominatim.
    """
    try:
        res = requests.get(
            f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json",
            headers={'User-Agent': 'MeteoApp'}
        )
        if res.ok:
            addr = res.json().get("address", {})
            ville  = addr.get("city") or addr.get("town") or addr.get("village") or ""
            region = addr.get("state") or ""
            pays   = addr.get("country") or ""
            return ville, region, pays
    except Exception as e:
        print("🌍 Exception géocodage :", e)
    return "", "", ""

def ajouter_donnees_excel(data):
    """
    Ajoute une ligne dans le fichier Excel avec les données météorologiques.
    """
    now        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    station_id = data['location'].get('tz_id', '')      # pas d'ID station METAR, on utilise le fuseau
    lat        = data['location']['lat']
    lon        = data['location']['lon']
    ville_loc, region, pays = get_adresse_reverse(lat, lon)
    cur        = data['current']
    description= cur['condition']['text']
    temp_c     = cur['temp_c']
    press_mb   = cur['pressure_mb']
    humidite   = cur['humidity']
    vent_ms    = cur['wind_kph'] / 3.6                  # conversion km/h → m/s
    precip_mm  = cur.get('precip_mm', 0.0)

    wb = load_workbook(fichier_excel)
    ws = wb.active
    ws.append([
        now, station_id, lat, lon,
        ville_loc, region, pays, description,
        temp_c, press_mb, humidite, vent_ms, precip_mm
    ])
    wb.save(fichier_excel)
    print(f"✅ Données ajoutées à {fichier_excel} à {now}")

if __name__ == "__main__":
    while True:
        try:
            r = requests.get(URL, timeout=10)
            if r.status_code == 200:
                ajouter_donnees_excel(r.json())
            else:
                print("⚠️ Erreur API météo :", r.status_code, r.text)
        except Exception as e:
            print("❌ Erreur :", e)
        # Pause de 10 minutes
        time.sleep(60)
