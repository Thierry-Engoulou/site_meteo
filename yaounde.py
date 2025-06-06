import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Votre clé WeatherAPI
API_KEY = "de330782dc16418397b172524250605"

# Option A : par nom de ville
CITY = "Yaoundé"

# Option B : directement par coordonnées de la station (lat,lon)
# Pour l’aéroport de Yaoundé‑Nsimalen (ICAO : FKYS) : 3.72N, 11.52E
COORD = "3.72,11.52"

# Choix de l’URL : utilisez CITY ou COORD
URL = f"http://api.weatherapi.com/v1/current.json?key={API_KEY}&q={CITY}&aqi=no"
# URL = f"http://api.weatherapi.com/v1/current.json?key={API_KEY}&q={COORD}&aqi=no"

fichier_excel = "meteo_yaounde.xlsx"

# Initialisation du fichier Excel
if not os.path.exists(fichier_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Météo"
    ws.append([
        "Date et Heure", "ID Station", "Latitude", "Longitude",
        "Ville", "Région", "Pays", "Description météo",
        "Température (°C)", "Pression (hPa)", "Humidité (%)",
        "Vitesse du vent (m/s)", "Précipitation (mm)"
    ])
    wb.save(fichier_excel)

def get_adresse_reverse(lat, lon):
    try:
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
        res = requests.get(url, headers={'User-Agent': 'MeteoApp'})
        if res.ok:
            addr = res.json().get("address", {})
            ville  = addr.get("city") or addr.get("town") or ""
            region = addr.get("state") or ""
            pays   = addr.get("country") or ""
            return ville, region, pays
    except:
        pass
    return "", "", ""

def ajouter_donnees_excel(data):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # WeatherAPI ne fournit pas d'ID station, on pourra utiliser le fuseau horaire
    station_id = data.get('location',{}).get('tz_id','')
    lat = data['location']['lat']
    lon = data['location']['lon']
    ville, region, pays = get_adresse_reverse(lat, lon)
    desc = data['current']['condition']['text']
    temp = data['current']['temp_c']
    pres = data['current']['pressure_mb']
    hum  = data['current']['humidity']
    wind = data['current']['wind_kph'] / 3.6
    precip = data['current'].get('precip_mm', 0.0)

    wb = load_workbook(fichier_excel)
    ws = wb.active
    ws.append([now, station_id, lat, lon,
               ville, region, pays, desc,
               temp, pres, hum, wind, precip])
    wb.save(fichier_excel)
    print(f"✅ Données ajoutées à {fichier_excel} à {now}")

if __name__ == "__main__":
    while True:
        try:
            r = requests.get(URL, timeout=10)
            if r.status_code == 200:
                ajouter_donnees_excel(r.json())
            else:
                print("⚠️ Erreur API météo :", r.status_code, r.text)
        except Exception as e:
            print("❌ Erreur :", e)
        time.sleep(60)  # toutes les 1 minutes
