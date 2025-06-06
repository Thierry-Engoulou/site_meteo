import time
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Votre clé WeatherAPI
API_KEY = "de330782dc16418397b172524250605"

# Liste de vos stations avec leur « id » ou leurs coordonnées
# Pour les aéroports, on met le code ICAO ; pour les autres, on met lat,lon
STATIONS = [
    {"name": "Douala–Nsimalen",     "q": "FKYS"},        # ICAO
    {"name": "Yaoundé–Nsimalen",    "q": "FKYS"},        # même ICAO
    {"name": "Maroua‑Salak",        "q": "10.58,14.18"},  # coord.
    {"name": "Garoua",              "q": "64860"},       # WMO (essayer)
    {"name": "Ngaoundéré",          "q": "64870"},       # WMO (essayer)
    {"name": "Bafoussam",           "q": "5.47,10.43"},   # coord.
    {"name": "Kribi",               "q": "2.94,9.91"},    # coord.
    {"name": "Mamfé",               "q": "5.78,9.30"},    # coord.
    {"name": "Tiko",                "q": "4.05,9.36"},    # coord.
    {"name": "Abong‑Mbang",         "q": "4.58,13.17"},   # coord.
    {"name": "Bamenda",             "q": "5.95,10.15"},   # coord.
    {"name": "Batouri",             "q": "4.48,14.37"},   # coord.
    {"name": "Betare‑Oya",          "q": "5.40,14.42"},   # coord.
    {"name": "Meiganga",            "q": "6.53,14.37"},   # coord.
]

# Nom du fichier Excel de sortie
FICHEIR_EXCEL = "meteo_cameroon_stations.xlsx"

# Prépare le fichier Excel s’il n’existe pas
if not os.path.exists(FICHEIR_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.title = "Météo Cameroun"
    ws.append([
        "Date et Heure", "Station", "ID/API-q",
        "Latitude", "Longitude",
        "Ville", "Région", "Pays",
        "Description", "Temp (°C)",
        "Pression (hPa)", "Humidité (%)",
        "Vent (m/s)", "Précip (mm)"
    ])
    wb.save(FICHEIR_EXCEL)

def fetch_station(q):
    """Interroge WeatherAPI pour le paramètre q (ICAO, WMO ou coord)."""
    url = f"http://api.weatherapi.com/v1/current.json?key={API_KEY}&q={q}&aqi=no"
    r = requests.get(url, timeout=10)
    if not r.ok:
        raise RuntimeError(f"Erreur API ({q}): {r.status_code} {r.text}")
    return r.json()

def get_adresse_reverse(lat, lon):
    """Reverse-geocode OSM pour ville/région/pays."""
    try:
        r = requests.get(
            f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json",
            headers={"User-Agent": "MeteoApp"}
        )
        addr = r.json().get("address", {})
        ville  = addr.get("city") or addr.get("town") or addr.get("village") or ""
        region = addr.get("state") or ""
        pays   = addr.get("country") or ""
        return ville, region, pays
    except:
        return "", "", ""

def enregistrer(d, station_name, q):
    """Ajoute une ligne dans le fichier Excel."""
    now    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    loc    = d["location"]
    cur    = d["current"]
    lat, lon = loc["lat"], loc["lon"]
    ville, region, pays = get_adresse_reverse(lat, lon)
    ligne = [
        now,
        station_name,
        q,
        lat,
        lon,
        ville,
        region,
        pays,
        cur["condition"]["text"],
        cur["temp_c"],
        cur["pressure_mb"],
        cur["humidity"],
        cur["wind_kph"] / 3.6,
        cur.get("precip_mm", 0.0)
    ]
    wb = load_workbook(FICHEIR_EXCEL)
    ws = wb.active
    ws.append(ligne)
    wb.save(FICHEIR_EXCEL)
    print(f"✅ {station_name} enregistré à {now}")

if __name__ == "__main__":
    while True:
        for st in STATIONS:
            try:
                data = fetch_station(st["q"])
                enregistrer(data, st["name"], st["q"])
            except Exception as e:
                print(f"⚠️ Erreur pour {st['name']} ({st['q']}):", e)
        # Attendre 1 minutes avant le prochain cycle
        time.sleep(60)
